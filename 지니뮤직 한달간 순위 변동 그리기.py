import requests
from bs4 import BeautifulSoup

from openpyxl import load_workbook

load_wb = load_workbook("mystock.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

dates = []

for i in range(1, 31):
    if i < 10:
        dates.append('2020030'+str(i))
    else:
        dates.append('202003'+str(i))

# print(dates)

j = 2
for date in dates:
    headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
    data = requests.get('https://www.genie.co.kr/chart/top200?ditc=D&ymd='+date+'&hh=23&rtm=N&pg=1', headers=headers)
    soup = BeautifulSoup(data.text, 'html.parser')

    songs = soup.select('#body-content > div.newest-list > div > table > tbody > tr')

    load_ws.cell(1, j, date) # 날짜 찍기

    # movies(tr들)의 반복문 돌리기
    i = 2
    for song in songs:
        rank = song.select_one('td.number').text[0:2].strip()
        title = song.select_one('td.info > a.title.ellipsis').text.strip()
        artist = song.select_one('td.info > a.artist.ellipsis').text

        load_ws.cell(i, j , title + '-' + artist)
        i += 1

    j += 1

load_wb.save("mystock.xlsx")